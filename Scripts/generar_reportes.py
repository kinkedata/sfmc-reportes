#!/usr/bin/env python3
"""
Generador de Reportes Diarios - Telcel Empresas (v1.5)
Lee CSV de SFMC y genera 11 Excel con protección de hojas + contraseña

REQUISITOS:
- openpyxl
- pandas
- msoffcrypto-tool (pip install msoffcrypto-tool) - OPCIONAL pero recomendado
"""

import os
import json
import csv
from datetime import datetime, timedelta
import random
import string
from pathlib import Path
import subprocess
import shutil

DIAS_ES = {
    'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
    'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
}

ORDEN_REGIONES = ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'MOM', 'BUI']

NOMBRES_REGION = {
    'R1': ('R1', 'la R1'), 'R2': ('R2', 'la R2'), 'R3': ('R3', 'la R3'),
    'R4': ('R4', 'la R4'), 'R5': ('R5', 'la R5'), 'R6': ('R6', 'la R6'),
    'R7': ('R7', 'la R7'), 'R8': ('R8', 'la R8'), 'R9': ('R9', 'la R9'),
    'MOM': ('Mobile Marketing', 'Mobile Marketing'),
    'BUI': ('Business Intelligence', 'Business Intelligence'),
}

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReporteLeads:
    def __init__(self, config_path='config.json', fechas_reporte=None):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)

        self.fecha_hoy = datetime.now()
        self.fecha_str = self.fecha_hoy.strftime('%d%m%Y')
        self.fecha_excel = self.fecha_hoy.strftime('%Y-%m-%d')
        self.dia_nombre = self.fecha_hoy.strftime('%A').capitalize()

        if fechas_reporte:
            self.fechas_reporte = set(fechas_reporte)
            if len(fechas_reporte) == 1:
                self.periodo_str = list(fechas_reporte)[0].strftime('%d-%m-%Y')
            else:
                d_min = min(fechas_reporte).strftime('%d-%m-%Y')
                d_max = max(fechas_reporte).strftime('%d-%m-%Y')
                self.periodo_str = f"{d_min} al {d_max}"
        else:
            self.fecha_ayer = self.fecha_hoy - timedelta(days=1)
            self.fechas_reporte = {self.fecha_ayer.date()}
            self.periodo_str = self.fecha_ayer.strftime('%d-%m-%Y')
        
        self.leads_por_region = {
            'R1': [], 'R2': [], 'R3': [], 'R4': [], 'R5': [],
            'R6': [], 'R7': [], 'R8': [], 'R9': [],
            'MOM': [], 'BUI': []
        }
        self._log = []
        self.contraseñas_generadas = {}
        self.archivos_generados = []

        self.crear_carpetas()
    
    def _print(self, msg=''):
        print(msg)
        self._log.append(str(msg))

    def crear_carpetas(self):
        """Crea carpetas de salida si no existen"""
        Path(self.config['rutas']['carpeta_salida']).mkdir(parents=True, exist_ok=True)
        Path(self.config['rutas']['carpeta_respaldo']).mkdir(parents=True, exist_ok=True)
    
    def leer_csv(self, ruta_csv):
        """Lee el CSV exportado de SFMC"""
        self._print(f"📖 Leyendo CSV: {ruta_csv}")

        if not os.path.exists(ruta_csv):
            raise FileNotFoundError(f"❌ No encontrado: {ruta_csv}")

        leads = []
        with open(ruta_csv, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                leads.append(row)

        self._print(f"✅ {len(leads)} registros cargados")
        return leads
    
    def filtrar_por_fecha(self, leads):
        """Filtra leads del día anterior - soporta múltiples formatos de fecha"""
        leads_filtrados = []
        formatos_fecha = [
            '%Y-%m-%d %H:%M:%S',      # 2026-02-05 18:53:43
            '%m/%d/%Y %I:%M:%S %p',   # 2/5/2026 6:53:43 PM (formato SFMC)
            '%m/%d/%Y %H:%M:%S',      # 2/5/2026 18:53:43
            '%d-%m-%Y %H:%M:%S',      # 05-02-2026 18:53:43
            '%m/%d/%Y %H:%M',         # 3/12/2026 22:49 (sin segundos)
            '%m/%d/%Y %I:%M %p',      # 3/12/2026 10:49 PM (sin segundos)
            '%Y-%m-%d %H:%M',         # 2026-02-05 18:53 (sin segundos)
        ]
        
        for lead in leads:
            try:
                fecha_str = lead['CreatedAt'].strip()
                fecha_creacion = None
                
                for formato in formatos_fecha:
                    try:
                        fecha_creacion = datetime.strptime(fecha_str, formato)
                        break
                    except ValueError:
                        continue
                
                if fecha_creacion and fecha_creacion.date() in self.fechas_reporte:
                    leads_filtrados.append(lead)
            except:
                continue

        self._print(f"📅 Período {self.periodo_str}: {len(leads_filtrados)} leads")
        return leads_filtrados
    
    def clasificar_leads(self, leads):
        """Clasifica leads por región y tipo de solución"""
        for lead in leads:
            region = lead.get('Region', '').upper()
            solucion = lead.get('SolucionInteres', '').upper()
            
            if 'MOBILE MARKETING' in solucion or solucion == 'MOM':
                self.leads_por_region['MOM'].append(lead)
            elif 'BUSINESS INTELLIGENCE' in solucion or solucion == 'BUI':
                self.leads_por_region['BUI'].append(lead)
            elif region in self.leads_por_region:
                self.leads_por_region[region].append(lead)
        
        self._print("\n📊 Distribución de Leads:")
        for region, leads_list in self.leads_por_region.items():
            if leads_list:
                self._print(f"  {region}: {len(leads_list)} leads")
    
    def generar_password(self, region):
        """Genera contraseña: R1TELXXX (3 caracteres aleatorios)"""
        caracteres = string.ascii_uppercase + string.digits
        random_chars = ''.join(random.choices(caracteres, k=3))
        return f"{region}TEL{random_chars}"
    
    def crear_excel(self, region, leads):
        """Crea Excel formateado para una región"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        header_fill = PatternFill(start_color="00529B", end_color="00529B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        columnas = self.config['columnas_reporte']
        for col_num, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = columna
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for row_num, lead in enumerate(leads, 2):
            for col_num, columna in enumerate(columnas, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = lead.get(columna, '')
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        for col_num, columna in enumerate(columnas, 1):
            ancho = max(len(str(columna)), 20)
            ws.column_dimensions[get_column_letter(col_num)].width = ancho

        return wb
    
    def _nombre_excel(self, region):
        """Genera el nombre de archivo Excel según la región"""
        if region in ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9']:
            return f"{region}-leads-calificados-x-dia-{region.lower()}-{self.fecha_excel}.xlsx"
        elif region == 'BUI':
            return f"BI-tel-leads-business-intellig-{self.fecha_excel}.xlsx"
        elif region == 'MOM':
            return f"MM-tel-leads-mobile-marketing-{self.fecha_excel}.xlsx"
        else:
            return f"{region}-leads-{self.fecha_excel}.xlsx"

    def guardar_excel_con_password(self, workbook, region, password):
        """Guarda Excel y lo protege con contraseña"""
        ruta = Path(self.config['rutas']['carpeta_salida']) / self._nombre_excel(region)
        
        # Guardar el Excel normalmente
        workbook.save(str(ruta))
        
        # Usar msoffcrypto-tool para proteger con contraseña al abrir
        try:
            # Renombrar el archivo original a temporal
            temp_ruta = str(ruta) + ".temp"
            shutil.move(str(ruta), temp_ruta)

            # Usar msoffcrypto-tool para encriptar (-e = encrypt)
            cmd = [
                'msoffcrypto-tool',
                '-e',
                '-p', password,
                temp_ruta,
                str(ruta)
            ]

            resultado = subprocess.run(cmd, capture_output=True, text=True)

            if resultado.returncode == 0:
                # Encriptación exitosa - eliminar temporal
                if os.path.exists(temp_ruta):
                    os.remove(temp_ruta)
                return ruta, password, True
            else:
                # Si falla, restaurar el archivo original sin contraseña
                if os.path.exists(temp_ruta):
                    shutil.move(temp_ruta, str(ruta))
                return ruta, password, False
        
        except FileNotFoundError:
            # msoffcrypto-tool no está instalado
            # Devolver el archivo sin encriptar
            return ruta, password, False
        except Exception as e:
            self._print(f"⚠️  Error protegiendo {region}: {e}")
            return ruta, password, False
    
    def generar_todos_reportes(self, ruta_csv):
        """Orquesta todo el proceso"""
        self._print(f"\n{'='*60}")
        self._print(f"🚀 GENERADOR DE REPORTES TELCEL EMPRESAS v1.5")
        self._print(f"{'='*60}")
        self._print(f"Fecha: {self.fecha_hoy.strftime('%d-%m-%Y %H:%M:%S')}\n")

        leads = self.leer_csv(ruta_csv)
        leads_validos = self.filtrar_por_fecha(leads)

        if not leads_validos:
            self._print(f"⚠️  No hay leads para el período {self.periodo_str}")
            return False

        self.clasificar_leads(leads_validos)

        archivos_generados = []
        contraseñas = {}

        self._print(f"\n📝 Generando archivos Excel con protección...\n")

        for region, leads_region in self.leads_por_region.items():
            if not leads_region:
                continue

            try:
                password = self.generar_password(region)
                contraseñas[region] = password

                wb = self.crear_excel(region, leads_region)
                ruta, pwd, protegido = self.guardar_excel_con_password(wb, region, password)

                estado = "🔒 PROTEGIDO" if protegido else "✅ CREADO"

                archivos_generados.append({
                    'region': region,
                    'archivo': ruta.name,
                    'ruta': str(ruta),
                    'leads': len(leads_region),
                    'password': pwd,
                    'protegido': protegido
                })

                self._print(f"✅ {region}: {ruta.name} ({len(leads_region)} leads) {estado} {pwd}")

            except Exception as e:
                self._print(f"❌ Error generando {region}: {str(e)}")

        self.contraseñas_generadas = contraseñas
        self.archivos_generados = archivos_generados
        self.crear_reporte_contraseñas(contraseñas)

        self._print(f"\n{'='*60}")
        self._print(f"✅ PROCESO COMPLETADO")
        self._print(f"{'='*60}\n")
        
        return True
    
    def _fecha_legible(self, d):
        """Devuelve 'Jueves 12' a partir de un objeto date."""
        return f"{DIAS_ES[d.strftime('%A')]} {d.day}"

    def _periodo_legible(self):
        """Devuelve el período en formato legible: 'Jueves 12' o 'Viernes 13 al Domingo 15'."""
        fechas = sorted(self.fechas_reporte)
        if len(fechas) == 1:
            return self._fecha_legible(fechas[0])
        return f"{self._fecha_legible(fechas[0])} al {self._fecha_legible(fechas[-1])}"

    def _generar_plantillas_correo(self, contraseñas):
        """Genera las 11 plantillas de correo para todas las regiones."""
        periodo = self._periodo_legible()
        lineas = [
            f"{'='*60}",
            f"PLANTILLAS DE CORREO - {periodo}",
            f"{'='*60}",
        ]

        for region in ORDEN_REGIONES:
            nombre_subject, nombre_mail = NOMBRES_REGION[region]
            lineas.append(f"\n{'─'*60}")
            lineas.append(f"  {nombre_subject}")
            lineas.append(f"{'─'*60}\n")

            lineas.append(f"Subject: Reporte de leads {nombre_subject} ({periodo})\n")

            if region in contraseñas:
                pwd = contraseñas[region]
                lineas.append(f"Buenos días equipo.")
                lineas.append(f"")
                lineas.append(f"Comparto el reporte de leads generados para {nombre_mail} correspondiente al período del {periodo}.")
                lineas.append(f"Contraseña: {pwd}")
                lineas.append(f"")
                lineas.append(f"Quedo pendiente.")
                lineas.append(f"Saludos.")
            else:
                lineas.append(f"Buenos días equipo.")
                lineas.append(f"")
                lineas.append(f"Les comento que no se han generado nuevos leads para {nombre_mail} del período del {periodo}.")
                lineas.append(f"")
                lineas.append(f"Quedo pendiente.")
                lineas.append(f"Saludos.")

        lineas.append(f"\n{'='*60}\n")
        return '\n'.join(lineas)

    def crear_reporte_contraseñas(self, contraseñas):
        """Crea archivo con contraseñas y log del proceso"""
        ruta = Path(self.config['rutas']['carpeta_salida']) / f"CONTRASEÑAS_{self.fecha_str}.txt"

        plantillas = self._generar_plantillas_correo(contraseñas)

        contraseñas_section = f"{'='*60}\nCONTRASEÑAS DIARIAS - {self.fecha_hoy.strftime('%d de %B de %Y')}\n{'='*60}\n\n"
        contraseñas_section += "IMPORTANTE: Los archivos Excel están protegidos con contraseña.\n"
        contraseñas_section += "Al abrir el archivo, se pedirá la contraseña que aparece aquí.\n\n"
        for region, password in contraseñas.items():
            contraseñas_section += f"{region:5s} :  {password}\n"
        contraseñas_section += f"\n{'='*60}\nINSTRUCCIONES:\n1. Abre el archivo Excel\n2. Se te pedirá la contraseña\n3. Ingresa la contraseña de arriba\n4. El archivo se abrirá correctamente\n\nNOTA: Estas contraseñas son válidas SOLO para hoy.\n"

        log_section = f"\n{'='*60}\nLOG DEL PROCESO\n{'='*60}\n"
        log_section += '\n'.join(self._log) + '\n'

        contenido = plantillas + contraseñas_section + log_section

        with open(ruta, 'w', encoding='utf-8') as f:
            f.write(contenido)

        self._print(f"🔐 Contraseñas guardadas en: {ruta.name}")
    

def main():
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

    try:
        generador = ReporteLeads(config_file)
        fecha_hoy = datetime.now().strftime('%d%m%Y')
        carpeta_csv = generador.config['rutas']['carpeta_csv']
        csv_entrada = os.path.join(carpeta_csv, f'leads_diarios_{fecha_hoy}.csv')
        generador.generar_todos_reportes(csv_entrada)
        
    except FileNotFoundError as e:
        print(f"❌ Error: {e}")
    except Exception as e:
        print(f"❌ Error inesperado: {e}")


if __name__ == "__main__":
    main()